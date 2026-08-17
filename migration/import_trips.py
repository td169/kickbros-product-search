#!/usr/bin/env python3
"""
One-time migration: import trip sheets from "Personal Shopping.xlsx" into the Supabase
`products` table. Run manually once; not wired into the app.

Usage:
    pip install openpyxl
    SUPABASE_URL=https://xxxxx.supabase.co SUPABASE_ANON_KEY=eyJ... \
        python3 import_trips.py "/path/to/Personal Shopping.xlsx" [--dry-run]

--dry-run parses and prints the summary without writing anything to Supabase — use this
first to sanity-check the brand inference and row counts before actually importing.
"""

import json
import os
import re
import sys
import urllib.request
from collections import defaultdict

import openpyxl

# Sheets that actually contain trip purchase data, in the "Name / EU Retail / After Tax /
# UK Retail / After Tax / Sale Price / Profit / Quantity / Total Profit [/ Total Spend]"
# format. Deliberately excludes:
#   - "Sheet1", "GOYARD", "Sheet2": an older, differently-structured (RRP/RRP AFTER/...)
#     draft layout that predates the trip-sheet format — not the schema this script maps.
#   - "Copy of PARIS JAN 2025", "Copy of PARIS JULY 2025": exact duplicates of the sheets
#     without "Copy of " in the name (re-imports would double-count).
#   - "PARIS BUDGET 2025": trip cost planning, not actual purchases.
#   - "Marketing": unrelated (networking contacts), not shopping data at all.
TRIP_SHEETS = [
    'PARIS JAN 2025',
    'PARIS FEB 2026',
    'PARIS JUNE 2025',
    'PARIS JULY 2025',
    'PARIS DECEMBER 2025',
    'PARIS JANUARY 2026',
    'SALES MISSED PARIS MID JULY 202',
    '1st time',
]

# Sub-section markers that can appear above a second (or third) "Name / EU Retail / ..."
# header within the same sheet. STOCK_MARKERS flip status to 'stock' for the rows below them;
# SKIP_MARKERS mark a hypothetical/what-if table (not real purchases) to skip entirely.
STOCK_MARKERS = ['stock only']
SKIP_MARKERS = ["what i could", "what i couldve", "what i could've"]

BRAND_KEYWORDS = [
    (r'\bdior\b', 'Dior'),
    (r'\blv\b|louis\s*vuitton', 'Louis Vuitton'),
    (r'\bgoyard\b', 'Goyard'),
    (r'\bgucci\b', 'Gucci'),
    (r'\bmoncler\b', 'Moncler'),
    (r'\bbalenciaga\b', 'Balenciaga'),
    (r'\bchanel\b', 'Chanel'),
    (r'herm[eè]s', 'Hermes'),
    (r'loro\s*pian[ao]', 'Loro Piana'),
    (r'tom\s*ford', 'Tom Ford'),
    (r'\bprada\b', 'Prada'),
]


def infer_brand(name):
    lower = name.lower()
    for pattern, brand in BRAND_KEYWORDS:
        if re.search(pattern, lower):
            return brand
    return None


def is_header_row(row):
    # Header rows are identified by "EU Retail" in the second column — the first column is
    # sometimes "Name", sometimes blank, so that's not a reliable anchor on its own.
    return len(row) > 1 and isinstance(row[1], str) and row[1].strip().lower() == 'eu retail'


def clean_num(v):
    if v is None or isinstance(v, str):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_sheet(ws, sheet_name):
    rows = list(ws.iter_rows(values_only=True))
    default_status = 'missed_sale' if 'sales missed' in sheet_name.lower() else 'sold'
    parsed = []
    unmatched_brands = []

    i = 0
    while i < len(rows):
        row = rows[i]
        if is_header_row(row):
            # Look a few rows back for a section-label row (a single text cell, no numbers)
            # to decide this table's status, or to skip it entirely if it's hypothetical.
            status = default_status
            skip_table = False
            for back in range(1, 4):
                idx = i - back
                if idx < 0:
                    break
                label_row = rows[idx]
                cells = [c for c in label_row if c is not None]
                if len(cells) == 1 and isinstance(cells[0], str):
                    label = cells[0].strip().lower()
                    if any(m in label for m in STOCK_MARKERS):
                        status = 'stock'
                    if any(m in label for m in SKIP_MARKERS):
                        skip_table = True
                    break
                elif cells:
                    break  # hit a non-label (data/expenses) row — stop looking back

            j = i + 1
            while j < len(rows):
                r = rows[j]
                name = r[0] if len(r) > 0 else None
                if not name or not str(name).strip():
                    break  # blank-name template row — end of this table
                if not skip_table:
                    name = str(name).strip()
                    eu_retail = clean_num(r[1] if len(r) > 1 else None)
                    eu_after_tax = clean_num(r[2] if len(r) > 2 else None)
                    uk_retail = clean_num(r[3] if len(r) > 3 else None)
                    uk_after_tax = clean_num(r[4] if len(r) > 4 else None)
                    sale_price = clean_num(r[5] if len(r) > 5 else None)
                    profit = clean_num(r[6] if len(r) > 6 else None)
                    quantity = clean_num(r[7] if len(r) > 7 else None) or 1
                    total_profit = clean_num(r[8] if len(r) > 8 else None)
                    total_spend = clean_num(r[9] if len(r) > 9 else None)
                    if total_profit is None and profit is not None:
                        total_profit = profit * quantity
                    if total_spend is None and eu_retail is not None:
                        total_spend = eu_retail * quantity

                    brand = infer_brand(name)
                    if not brand:
                        unmatched_brands.append((sheet_name, name))

                    parsed.append({
                        'brand': brand,
                        'product_name': name,
                        'trip_label': sheet_name,
                        'eu_retail': eu_retail,
                        'eu_after_tax': eu_after_tax,
                        'uk_retail': uk_retail,
                        'uk_after_tax': uk_after_tax,
                        'sale_price': sale_price,
                        'profit': profit,
                        'quantity': int(quantity),
                        'total_profit': total_profit,
                        'total_spend': total_spend,
                        'status': status,
                    })
                j += 1
            i = j
        else:
            i += 1

    return parsed, unmatched_brands


def insert_rows(url, key, rows):
    endpoint = url.rstrip('/') + '/rest/v1/products'
    body = json.dumps(rows).encode('utf-8')
    req = urllib.request.Request(endpoint, data=body, method='POST', headers={
        'apikey': key,
        'Authorization': f'Bearer {key}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal',
    })
    with urllib.request.urlopen(req) as res:
        return res.status


def main():
    args = sys.argv[1:]
    dry_run = '--dry-run' in args
    args = [a for a in args if a != '--dry-run']
    if not args:
        print(__doc__)
        sys.exit(1)
    xlsx_path = args[0]

    url = os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_ANON_KEY')
    if not dry_run and (not url or not key):
        print('Set SUPABASE_URL and SUPABASE_ANON_KEY env vars (or pass --dry-run to skip writing).')
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_rows = []
    all_unmatched = []
    per_trip_counts = defaultdict(int)
    per_trip_status = defaultdict(lambda: defaultdict(int))

    for sheet_name in TRIP_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f'!! Sheet not found, skipping: {sheet_name}')
            continue
        rows, unmatched = parse_sheet(wb[sheet_name], sheet_name)
        all_rows.extend(rows)
        all_unmatched.extend(unmatched)
        per_trip_counts[sheet_name] = len(rows)
        for r in rows:
            per_trip_status[sheet_name][r['status']] += 1

    print(f'\n{"=" * 60}\nParsed {len(all_rows)} rows total\n{"=" * 60}')
    for trip, count in per_trip_counts.items():
        status_bits = ', '.join(f'{s}={n}' for s, n in per_trip_status[trip].items())
        print(f'  {trip}: {count} rows ({status_bits})')

    if all_unmatched:
        print(f'\n!! {len(all_unmatched)} rows with no brand match (left blank for manual review):')
        for trip, name in all_unmatched:
            print(f'    [{trip}] {name!r}')

    if dry_run:
        print('\n--dry-run: nothing written to Supabase.')
        return

    print(f'\nInserting {len(all_rows)} rows into Supabase...')
    # Batch in chunks to keep individual request bodies reasonable.
    CHUNK = 50
    for i in range(0, len(all_rows), CHUNK):
        chunk = all_rows[i:i + CHUNK]
        status = insert_rows(url, key, chunk)
        print(f'  rows {i+1}-{i+len(chunk)}: HTTP {status}')

    print('Done.')


if __name__ == '__main__':
    main()
